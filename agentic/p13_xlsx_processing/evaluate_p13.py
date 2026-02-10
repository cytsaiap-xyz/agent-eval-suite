#!/usr/bin/env python3
"""Evaluation script for P13: Excel Processing"""

import json
import sys
from pathlib import Path


def evaluate(base_path: str = ".") -> dict:
    base = Path(base_path)
    results = {"scores": {}, "details": [], "total_score": 0, "max_score": 100}

    xlsx_path = base / "financial_data.xlsx"
    if not xlsx_path.exists():
        results["details"].append("✗ financial_data.xlsx not found")
        return results

    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(xlsx_path))
        sheet_names = wb.sheetnames

        # Check for Data_Quality sheet
        if "Data_Quality" in sheet_names:
            results["scores"]["data_quality"] = 15
            results["details"].append("✓ Data_Quality sheet created")
        else:
            results["scores"]["data_quality"] = 0
            results["details"].append("✗ Data_Quality sheet not found")

        # Check for calculated columns in Sales
        ws_sales = wb["Sales"]
        headers = [cell.value for cell in ws_sales[1]]
        expected_cols = ["Revenue", "Revenue_USD", "Quarter"]
        found_cols = [c for c in expected_cols if c in headers]

        if len(found_cols) >= 2:
            results["scores"]["calculations"] = 20
            results["details"].append(f"✓ Calculated columns added: {found_cols}")
        elif len(found_cols) > 0:
            results["scores"]["calculations"] = 10
            results["details"].append(f"◐ Some calculated columns: {found_cols}")
        else:
            results["scores"]["calculations"] = 0
            results["details"].append("✗ No calculated columns found")

        # Check for Pivot_Analysis sheet
        if "Pivot_Analysis" in sheet_names:
            results["scores"]["pivot"] = 15
            results["details"].append("✓ Pivot_Analysis sheet created")
        else:
            results["scores"]["pivot"] = 0
            results["details"].append("✗ Pivot_Analysis sheet not found")

        # Check for Dashboard sheet
        if "Dashboard" in sheet_names:
            results["scores"]["dashboard"] = 25
            results["details"].append("✓ Dashboard sheet created")
        else:
            results["scores"]["dashboard"] = 0
            results["details"].append("✗ Dashboard sheet not found")

        # Check for charts (basic check)
        charts_found = 0
        for sheet in wb.worksheets:
            if hasattr(sheet, '_charts'):
                charts_found += len(sheet._charts)

        if charts_found >= 3:
            results["scores"]["charts"] = 15
            results["details"].append(f"✓ {charts_found} charts found")
        elif charts_found > 0:
            results["scores"]["charts"] = charts_found * 3
            results["details"].append(f"◐ Only {charts_found} charts")
        else:
            results["scores"]["charts"] = 0
            results["details"].append("✗ No charts found")

    except ImportError:
        results["details"].append("⚠ openpyxl not installed, limited validation")
        results["scores"]["file_exists"] = 30
    except Exception as e:
        results["details"].append(f"✗ Error reading xlsx: {e}")

    # Check for analysis_summary.json
    json_path = base / "analysis_summary.json"
    if json_path.exists():
        results["scores"]["summary_json"] = 10
        results["details"].append("✓ analysis_summary.json created")
    else:
        results["scores"]["summary_json"] = 0
        results["details"].append("✗ analysis_summary.json not found")

    results["total_score"] = sum(results["scores"].values())
    return results


def main():
    base_path = sys.argv[1] if len(sys.argv) > 1 else "."
    print("=" * 60)
    print("Problem 13: Excel Processing - Evaluation")
    print("=" * 60)

    results = evaluate(base_path)

    for detail in results["details"]:
        print(f"  {detail}")

    print(f"\nTotal Score: {results['total_score']}/{results['max_score']}")

    with open(Path(base_path) / "evaluation_results.json", "w") as f:
        json.dump(results, f, indent=2)


if __name__ == "__main__":
    main()
