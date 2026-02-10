#!/usr/bin/env python3
"""
Generate the sample Excel workbook for P13.
Requires: pip install openpyxl
"""

from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
import random


def create_sample_workbook():
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Sheet 1: Sales
    ws_sales = wb.create_sheet("Sales")
    sales_headers = ["Order_ID", "Date", "Customer_ID", "Product_ID", "Region_ID",
                     "Quantity", "Unit_Price", "Currency", "Discount_Pct"]
    ws_sales.append(sales_headers)

    products = ["P001", "P002", "P003", "P004", "P005", "P006", "P007", "P008"]
    regions = ["R01", "R02", "R03", "R04", "R05"]
    currencies = ["USD", "EUR", "GBP", "JPY"]

    # Generate 500 sales records
    base_date = datetime(2024, 1, 1)
    for i in range(500):
        order_id = f"ORD{10001 + i}"
        date = base_date + timedelta(days=random.randint(0, 364))
        customer_id = f"C{random.randint(1001, 1100)}"
        product_id = random.choice(products)
        region_id = random.choice(regions)
        quantity = random.randint(1, 50)
        unit_price = round(random.uniform(10, 500), 2)
        currency = random.choice(currencies)
        discount = random.choice([0, 0, 0, 5, 10, 15, 20])

        # Add some data quality issues
        if i == 50:  # Duplicate
            ws_sales.append(["ORD10050", date, customer_id, product_id, region_id,
                           quantity, unit_price, currency, discount])
        if i == 100:  # Missing value
            ws_sales.append([order_id, date, None, product_id, region_id,
                           quantity, unit_price, currency, discount])
        if i == 150:  # Negative quantity (anomaly)
            ws_sales.append([order_id, date, customer_id, product_id, region_id,
                           -5, unit_price, currency, discount])

        ws_sales.append([order_id, date, customer_id, product_id, region_id,
                        quantity, unit_price, currency, discount])

    # Sheet 2: Products
    ws_products = wb.create_sheet("Products")
    ws_products.append(["Product_ID", "Product_Name", "Category", "Cost", "Launch_Date"])
    products_data = [
        ["P001", "Widget Pro", "Electronics", 45.00, datetime(2022, 1, 15)],
        ["P002", "Gadget Plus", "Electronics", 120.00, datetime(2022, 6, 1)],
        ["P003", "Tool Master", "Tools", 35.00, datetime(2021, 3, 10)],
        ["P004", "Smart Device", "Electronics", 200.00, datetime(2023, 2, 20)],
        ["P005", "Basic Kit", "Accessories", 15.00, datetime(2020, 8, 5)],
        ["P006", "Premium Set", "Accessories", 75.00, datetime(2023, 4, 12)],
        ["P007", "Industrial Tool", "Tools", 250.00, datetime(2022, 11, 30)],
        ["P008", "Consumer Pack", "Accessories", 25.00, datetime(2024, 1, 1)],
    ]
    for row in products_data:
        ws_products.append(row)

    # Sheet 3: Regions
    ws_regions = wb.create_sheet("Regions")
    ws_regions.append(["Region_ID", "Region_Name", "Country", "Manager"])
    regions_data = [
        ["R01", "North America", "USA", "John Smith"],
        ["R02", "Western Europe", "Germany", "Anna Mueller"],
        ["R03", "APAC", "Singapore", "Wei Chen"],
        ["R04", "Latin America", "Brazil", "Carlos Silva"],
        ["R05", "Eastern Europe", "Poland", "Marta Kowalski"],
    ]
    for row in regions_data:
        ws_regions.append(row)

    # Sheet 4: Targets
    ws_targets = wb.create_sheet("Targets")
    ws_targets.append(["Month", "Region_ID", "Target_Revenue", "Target_Orders"])
    months = ["2024-01", "2024-02", "2024-03", "2024-04", "2024-05", "2024-06",
              "2024-07", "2024-08", "2024-09", "2024-10", "2024-11", "2024-12"]
    for month in months:
        for region in regions:
            target_rev = random.randint(50000, 150000)
            target_orders = random.randint(30, 100)
            ws_targets.append([month, region, target_rev, target_orders])

    # Sheet 5: Exchange Rates
    ws_fx = wb.create_sheet("Exchange_Rates")
    ws_fx.append(["Currency", "To_USD_Rate", "Effective_Date"])
    fx_data = [
        ["USD", 1.00, datetime(2024, 1, 1)],
        ["EUR", 1.10, datetime(2024, 1, 1)],
        ["GBP", 1.27, datetime(2024, 1, 1)],
        ["JPY", 0.0067, datetime(2024, 1, 1)],
    ]
    for row in fx_data:
        ws_fx.append(row)

    # Save
    wb.save("financial_data.xlsx")
    print("Created financial_data.xlsx with 5 sheets")
    print("- Sales: 500+ records with some data quality issues")
    print("- Products: 8 products")
    print("- Regions: 5 regions")
    print("- Targets: Monthly targets by region")
    print("- Exchange_Rates: 4 currencies")


if __name__ == "__main__":
    create_sample_workbook()
